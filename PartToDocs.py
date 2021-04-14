import os, pyperclip, openpyxl
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties

srcFolder = pyperclip.paste()
os.chdir(srcFolder)

SWparts = []
partNames =[]

# Find valid folder and add them to a list of parts
def findValidPart():
    for file in os.listdir(srcFolder):
        if file.upper().endswith('.SLDPRT') or file.upper().endswith('.SLDASM'):
            if file.startswith('~$'):
                break
            SWparts.append(file)

# Extract Part names from SW parts
def extractNames(SWparts):
    for part in SWparts:
        partNames.append(part[:12])
        

findValidPart()
extractNames(SWparts)
print('Solidworks parts in this folder: ')
print(partNames)


# Get File Number
def getFileNo(partNames):
    partName = partNames[1]
    fileNo = partName[:4]
    return fileNo
        
# Define fileNo
fileNo = getFileNo(partNames)
print('This File Number is: ' + '"' + fileNo + '"')

# Get Assemblies
assys = set()
def getAssyNos(partNames):
    for part in partNames:
        assys.add(part[:9])



getAssyNos(partNames)
assyList = list(assys)
assyList.sort()
print('The assemblies in this folder are: ')
print(assyList)


# load the XL file for this this fileNo

wb = openpyxl.load_workbook('../' + (fileNo + ' - Part Numbers.xlsx'))

wsList = list(wb.sheetnames)

XlFileName = '../' + fileNo + ' - Part Numbers.xlsx'

#print(wsList.index(assyList[0]))

partsSetInSheet = set()

# Make list of drawings in a Sheet
def getPartsInSheet(sheetName):
    ws = wb[sheetName]
    for row in ws.iter_rows(min_row=5, max_col=1, max_row=99, values_only=True):
        for cell in row:
            if cell != None:
                partsSetInSheet.add(cell)
                

# Make a list of parts from each assy
for assy in assyList:
    assyName = assy
    assy = set()
    for part in partNames:
        if assyName == part[:9]:
            assy.add(part)
    print('Parts in assembly ' + assyName + ' are:')
    assyParts = list(assy)
    assyParts.sort()
    print(assyParts)        
    for worksheet in wsList:
        for ass in assyList:
            if worksheet == ass:
                getPartsInSheet(ass)
                listOfPartsInSheet = list(partsSetInSheet)
                listOfPartsInSheet.sort()
                #Winging it here - trying to compare each part
                for eachPart in listOfPartsInSheet:
                    for otherPart in assyParts:
                        if eachPart == otherPart:
                            continue
                        else:
                            ws = wb[ass]
                            for cell in ws.iter_rows(min_row=5, max_col=1, max_row=99, values_only=True):
                                if cell == None:
                                    cell.value = otherPart
                                    wb.save(XlFileName)
                                    break
                getPartsInSheet(ass)
                listOfPartsInSheet = list(partsSetInSheet)
                listOfPartsInSheet.sort()
                print('Update list of parts in the sheet are: ')
                print(listOfPartsInSheet)







print('The parts that are already documented are: ')
print(listOfPartsInSheet)